import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

class ExcelProcessor:
    def __init__(self, file_path, cutoff_10th, cutoff_12th, cutoff_btech_cgpa,
                 cutoff_live_kt, cutoff_drop, cutoff_gap):
        self.file_path = file_path
        self.cutoff_10th = cutoff_10th
        self.cutoff_12th = cutoff_12th
        self.cutoff_btech_cgpa = cutoff_btech_cgpa
        self.cutoff_live_kt = cutoff_live_kt
        self.cutoff_drop = cutoff_drop
        self.cutoff_gap = cutoff_gap
        self.df = pd.read_excel(self.file_path, sheet_name='Form responses 1')

    def clean_data(self):
        self.df.iloc[:, 0] = self.df.iloc[:, 0].astype(str).str.replace(' ', '')
        self.df.drop_duplicates(subset=self.df.columns[0], inplace=True)
        self.df.sort_values(by=self.df.columns[0], inplace=True)

    def check_eligibility_Percentage(self, column, cutoff, removed_df):
        removed = self.df[self.df[column] < cutoff]
        remaining = self.df[self.df[column] >= cutoff]
        removed_df = pd.concat([removed_df, removed])
        return remaining, removed_df

    def check_eligibility_Diff(self, column, cutoff, removed_df):
        removed = self.df[self.df[column] > cutoff]
        remaining = self.df[self.df[column] <= cutoff]
        removed_df = pd.concat([removed_df, removed])
        return remaining, removed_df

    def check_year_format(self, column_name):
        invalid_year_format = ~self.df[column_name].astype(str).str.match(r'^\d{4}$')
        if invalid_year_format.any():
            print(f"Invalid {column_name} format at rows:")
            print(self.df[invalid_year_format])

    def check_data(self):
        removed_df = pd.DataFrame(columns=self.df.columns)
        self.df, removed_df = self.check_eligibility_Percentage('10th Percentage', self.cutoff_10th, removed_df)
        self.check_year_format('10th Year of Passing')
        self.df, removed_df = self.check_eligibility_Percentage('12th/ Diploma Percentage', self.cutoff_12th, removed_df)
        self.check_year_format('12th/ Diploma Year of Passing')
        self.df, removed_df = self.check_eligibility_Percentage('BTech CGPA', self.cutoff_btech_cgpa, removed_df)
        self.df, removed_df = self.check_eligibility_Diff('Live KT', self.cutoff_live_kt, removed_df)
        self.df, removed_df = self.check_eligibility_Diff('Drop', self.cutoff_drop, removed_df)
        self.df, removed_df = self.check_eligibility_Diff('Gap', self.cutoff_gap, removed_df)
        return removed_df

    def remove_columns(self):
        columns_to_remove = ['Timestamp', 'Email address', 'College Roll No', 'Job Role', 'Minor Course', 'Live KT',
                             'Dead KT', 'Drop', 'Gap', 'Are you placed?', 'Company Name 1', 'Company Name 2',
                             'Statement of Acknowledgement']
        columns_to_remove = [col for col in columns_to_remove if col in self.df.columns]
        self.df.drop(columns_to_remove, axis=1, inplace=True)

    def basic_editing(self):
        self.df.iloc[:, 1] = self.df.iloc[:, 1].str.strip()

    def remove_unwanted_data(self, percentage):
        if percentage:
            self.df.drop('BTech CGPA', axis=1, inplace=True)
        else:
            self.df.drop('BTech Percentage', axis=1, inplace=True)

    def format_sheet(self):
        self.df.style.set_properties(**{
            'font-family': 'Arial',
            'font-size': '10pt',
            'text-align': 'center',
        }).set_table_styles(
            [{'selector': 'th', 'props': [('font-weight', 'bold')]}]
        )

    def rename_sheet(self, sheet_name):
        self.df.rename(columns={'Full Name': 'Name'}, inplace=True)

    def fill_empty_cells_with_na(self):
        self.df.fillna('NA', inplace=True)

    def sort_columns(self, preferred_order):
        existing_columns = list(self.df.columns)
        ordered_columns = [col for col in preferred_order if col in existing_columns] + \
                          [col for col in existing_columns if col not in preferred_order]
        self.df = self.df[ordered_columns]
        self.df = self.df.sort_values(by='Full Name', ascending=True)

    def add_serial_column(self):
        self.df.insert(0, 'Sr No', range(1, len(self.df) + 1))

    def save_data(self, output_file, removed_df, sheet_name='RAIT'):
        # Save the DataFrame to an Excel file with hyperlinks preserved
        full_otPath = os.path.join(os.getcwd(), output_file)

        with pd.ExcelWriter(full_otPath, engine='openpyxl') as writer:
            self.df.to_excel(writer, sheet_name=sheet_name, index=False)
            removed_df.to_excel(writer, sheet_name='removed', index=False)

        wb = load_workbook(full_otPath)
        ws = wb[sheet_name]

        # Reapply hyperlinks to the Resume column
        resume_col_idx = self.df.columns.get_loc('Resume') + 1
        for idx, url in enumerate(self.df['Resume'], start=2):  # Data starts from row 2
            if pd.notna(url):
                ws.cell(row=idx, column=resume_col_idx).hyperlink = url
                ws.cell(row=idx, column=resume_col_idx).style = "Hyperlink"

        wb.save(full_otPath)

    def adjust_column_widths(self, input_file, output_file):
        wb = load_workbook(input_file)
        ws = wb.active

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

                    cell.alignment = Alignment(horizontal='center')

                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='center')

                except TypeError:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

        return wb
